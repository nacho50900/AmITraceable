#!/usr/bin/env python3
"""
Recalcula y (con --apply) escribe STUDIES_DISTRIBUTION en ine_reference.py
a partir de ficheros Excel del Ministerio de Ciencia/Universidades --
NO es una API en vivo como el resto de tablas de update_ine_reference.py
(no se ha encontrado ninguna para esta fuente concreta, ver el historial
de investigación en el docstring de ese otro script), así que este toma
como entrada ficheros descargados A MANO, no hace ninguna llamada de red.

Reemplaza a `compute_studies_distribution.py` (que solo IMPRIMÍA el
resultado para copiar/pegar a mano) -- a partir de esta sesión el script
también ESCRIBE el fichero, igual que ya hacen todos los `_apply_*` de
update_ine_reference.py, reutilizando exactamente esas mismas funciones
(`_locate_block`, `_apply_float_block`, `_confirm_and_write`,
`_update_last_verified`) en vez de duplicar esa lógica.

METODOLOGÍA (dos pasos, para evitar dos problemas distintos del método
anterior -- ver discusión con Nacho en esta sesión):

  1. TOTAL POR RAMA (5 categorías amplias: Ciencias de la Salud,
     Ciencias Sociales y Jurídicas, Ingeniería y Arquitectura, Ciencias,
     Artes y Humanidades), sumando EGRESADOS (graduados) año a año desde
     el fichero histórico "Total SUE" (todos los niveles, desde 1984 si
     está disponible) -- EGRESADOS y no MATRICULADOS a propósito: cada
     persona aparece una sola vez, en el año que se titula, así que sumar
     egresados de muchos años SÍ da un recuento acumulado válido. Sumar
     MATRICULADOS de una serie larga, en cambio, cuenta a la misma
     persona una vez por cada año que estuvo matriculada en la misma
     carrera (3-6 veces) -- con 40 años de histórico eso infla el total
     muy por encima de la población real (la señal de alarma que dio
     Nacho: "hemos pasado de 46 millones de personas a 100").
  2. REPARTO DENTRO DE CADA RAMA, con el detalle real por titulación
     2015-2023/2024 que ya tenemos (Matriculados/Egresados por carrera
     concreta) -- asumiendo que el reparto reciente entre carreras de una
     misma rama aproxima el reparto histórico (aproximación explícita,
     no medición directa).

  total_estimado[carrera] = total_historico_egresados[rama de esa carrera]
                             x reparto_reciente[carrera dentro de su rama]

  proporcion_final[carrera] = total_estimado[carrera] / poblacion_25_64

`poblacion_25_64` se calcula de verdad (no se asume un "40%"): a partir
de TOTAL_POPULATION_ES x suma de AGE_DISTRIBUTION_5Y para los tramos
25-29 a 60-64, ambos ya en ine_reference.py -- da ~27,7M, más fundamentado
que la estimación anterior.

SI NO SE DA EL FICHERO HISTÓRICO (--egresados-historico-rama): el script
sigue funcionando con el método anterior (reparto reciente x 40% de
población con estudios superiores, ver versión previa de este cálculo en
git) como fallback explícito, avisando claramente de que está usando el
método menos fundamentado.

ADEMÁS (pedido explícito de Nacho: "sería útil tener las tablas
completas... si sabemos que uno es de la Universidad de Oviedo y su
año"): exporta el detalle COMPLETO por universidad/titulación/año a un
JSON aparte (backend/app/data/studies_by_university.json por defecto),
para poder usarlo en el futuro en análisis más finos que sepan la
universidad y el año concretos de alguien, no solo el agregado nacional
que usa STUDIES_DISTRIBUTION.

USO:
    python3 update_studies_distribution.py \\
        --matriculados MatriculadosTitulacion2015_2024.xlsx \\
        --egresados EgresadosTitulacion2015_2023.xlsx \\
        [--egresados-historico-rama HIS_Egr_TotalSUE_Rama_Univ.xlsx] \\
        [--apply] [--yes]

Requiere: pip install openpyxl httpx --break-system-packages
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Falta el paquete 'openpyxl'. Instálalo con:")
    print("    pip install openpyxl --break-system-packages")
    sys.exit(1)

# Reutiliza los helpers de escritura de fichero de update_ine_reference.py
# (mismo directorio) en vez de duplicarlos -- _locate_block/_apply_float_block/
# _confirm_and_write/_update_last_verified ya están probados contra las
# otras 6 tablas de ese script.
sys.path.insert(0, str(Path(__file__).parent))
import update_ine_reference as _uir  # noqa: E402

_JSON_EXPORT_PATH = Path(__file__).parent.parent / "app" / "data" / "studies_by_university.json"
_CACHE_DIR = Path(__file__).parent / ".cache_studies"

# --- Ficheros SIN año en la ruta de descarga (sistema /dam/jcr:...) ---
# El identificador `jcr:<hash>` es lo que de verdad decide qué fichero se
# sirve -- es un ID opaco de la biblioteca de activos (DAM) del
# Ministerio, SIN relación con el año. El "2015_2024" que aparece en el
# nombre del fichero es solo el nombre para "guardar como" del navegador;
# probar con otros años ahí (2015_2025, 2015_2026...) devolvería el
# MISMO fichero fijado a este hash, no uno más nuevo -- así que aquí NO
# se implementa un bucle de años, sería un teatro que no encuentra nada
# de verdad. Cuando el Ministerio publique una edición nueva bajo un hash
# distinto, hay que volver a buscarlo a mano y actualizar este diccionario
# (ver _INSTRUCCIONES_MANUALES más abajo para cómo).
_URLS_SIN_AÑO: dict[str, str] = {
    "matriculados": "https://www.ciencia.gob.es/dam/jcr:717ab000-0372-44e4-bec3-e49afcb2838b/MatriculadosTitulacion2015_2024.xlsx",
    "egresados": "https://www.ciencia.gob.es/dam/jcr:decc2024-82d0-4063-9202-9be6885bd34a/EgresadosTitulacion2015_2023.xlsx",
}

# --- Ficheros CON año en la ruta de descarga (sistema jaxiPx) ---
# A diferencia de los de arriba, aquí el año SÍ es parte real de la ruta
# (".../EEU_2024/Serie/TotalSUE/..."), porque el Ministerio publica una
# carpeta nueva cada curso bajo ese patrón -- confirmado con la hermana
# de matriculados de este mismo fichero (HIS_Mat_TotalSUE_Rama_Univ.xlsx,
# bajo EEU_2024). Aquí SÍ tiene sentido ir probando de más reciente a más
# antiguo, para que este script siga funcionando en cursos futuros sin
# tener que tocar el código -- ver `_descargar_con_fallback_de_año`.
#
# CONFIRMADO CONTRA LA RED REAL (Nacho, sesión posterior a cuando se
# escribió el bucle de plantillas múltiples): de las 4 combinaciones que
# se probaban antes, solo UNA funciona de verdad -- dominio
# estadisticas.ciencia.gob.es, formato .xlsx. Las otras 3 se han quitado
# del bucle activo para no perder tiempo con intentos que sabemos que
# van a fallar:
#   - estadisticas.universidades.gob.es (cualquier formato): siempre
#     'RemoteProtocolError: Server disconnected without sending a
#     response' en los 5 años probados -- corte activo de conexión, no
#     un 404, probablemente ese dominio no sirve así este fichero en
#     concreto (aunque sí sirve los de MATRICULADOS sin año, ver
#     _URLS_SIN_AÑO más arriba -- no es un bloqueo genérico del dominio).
#   - formato .csv/.px en estadisticas.ciencia.gob.es: 404 limpio (no
#     existe esa combinación de ruta).
# Si en el futuro esto deja de funcionar, antes de reactivar las otras
# combinaciones de abajo (quedan comentadas, no borradas) prueba primero
# a mano con --insecure si el fallo es de red/SSL o si de verdad ya no
# existe esa ruta.
_PLANTILLAS_URL_CON_AÑO: dict[str, list[str]] = {
    "egresados_historico_rama": [
        "https://estadisticas.ciencia.gob.es/jaxiPx/files/_px/es/xlsx/Universitaria/Alumnado/EEU_{año}/Serie/TotalSUE/l0/HIS_Egr_TotalSUE_Rama_Univ.xlsx",
        # Descartadas por no funcionar en la prueba real (ver comentario de arriba) -- se dejan aquí, comentadas, como referencia:
        # "https://estadisticas.universidades.gob.es/jaxiPx/files/_px/es/xlsx/Universitaria/Alumnado/EEU_{año}/Serie/TotalSUE/l0/HIS_Egr_TotalSUE_Rama_Univ.xlsx",
        # "https://estadisticas.universidades.gob.es/jaxiPx/files/_px/es/csv/Universitaria/Alumnado/EEU_{año}/Serie/TotalSUE/l0/HIS_Egr_TotalSUE_Rama_Univ.px",
        # "https://estadisticas.ciencia.gob.es/jaxiPx/files/_px/es/csv/Universitaria/Alumnado/EEU_{año}/Serie/TotalSUE/l0/HIS_Egr_TotalSUE_Rama_Univ.px",
    ],
}
_PRIMER_AÑO_CONOCIDO = 2023  # el más antiguo que se ha visto funcionar de verdad -- tope inferior del bucle, no bajar de aquí sin motivo
# El Ministerio publica los datos del curso que ACABA de terminar, no por
# adelantado -- así que no hay motivo para mirar muy por delante del año
# actual. El margen de 1 (no 0) es solo por una ambigüedad real que sí se
# ha visto en los ejemplos encontrados: no está confirmado si la carpeta
# se nombra por el año de INICIO o de FIN del curso (p. ej. si "curso
# 2025-2026" cae bajo "EEU_2025" o "EEU_2026") -- con margen 1 se cubren
# las dos convenciones sin tener que decidir cuál es la correcta.
_MARGEN_AÑOS_FUTUROS = 1
_PAUSA_ENTRE_INTENTOS_SEG = 1.5  # entre cada intento del bucle de año x plantilla -- por si el corte de conexión visto era anti-bot por peticiones demasiado seguidas

# Instrucciones de navegación manual, para cuando la descarga automática
# falle -- MUY explícitas a propósito (ruta completa clic a clic, no solo
# "búscalo en la web"), porque este es el punto de fallo más probable de
# todo el script: fuente sin API, con URLs que pueden cambiar de un curso
# a otro, y sin manera de que Claude verifique esto último por su cuenta
# (el dominio bloquea el acceso automatizado de Claude por robots.txt).
_INSTRUCCIONES_MANUALES: dict[str, str] = {
    "matriculados": (
        "1. Ir a https://www.ciencia.gob.es/Ministerio/Estadisticas/SIIU/UCT.html\n"
        "   (o buscar en Google: site:ciencia.gob.es Matriculados Titulacion xlsx)\n"
        "2. Buscar el enlace 'Matriculados' (Excel) en la sección de\n"
        "   Estadística de Estudiantes Universitarios (EEU).\n"
        "3. Pasar la ruta del .xlsx descargado con --matriculados\n"
        "4. (opcional, para que las próximas ejecuciones ya no necesiten\n"
        "   --matriculados) actualizar _URLS_SIN_AÑO['matriculados'] en\n"
        "   este script con la URL nueva -- el hash 'jcr:...' habrá\n"
        "   cambiado si el Ministerio publicó un activo nuevo."
    ),
    "egresados": (
        "1. Mismo sitio que 'matriculados' arriba, enlace 'Egresados' en vez\n"
        "   de 'Matriculados'.\n"
        "2. Pasar la ruta del .xlsx descargado con --egresados\n"
        "3. (opcional) actualizar _URLS_SIN_AÑO['egresados'] en este script."
    ),
    "egresados_historico_rama": (
        "1. Ir a https://estadisticas.ciencia.gob.es/jaxiPx/Tabla.htm?"
        "path=/Universitaria/Alumnado/EEU_2025/Serie/TotalSUE/l0/&"
        "file=HIS_Egr_TotalSUE_Rama_Univ.px&L=0 (cambia el año de la ruta\n"
        "   si hace falta -- este SÍ sigue un patrón por año, ver\n"
        "   _PLANTILLA_URL_CON_AÑO)\n"
        "2. Buscar el botón de descarga/exportar de esa página (icono de\n"
        "   Excel o 'Descargar') y guardar como .xlsx.\n"
        "3. Pasar la ruta del .xlsx descargado con --egresados-historico-rama\n"
        "   (si no se consigue, el script sigue funcionando con el método\n"
        "   de respaldo -- este fichero mejora la precisión pero no es\n"
        "   obligatorio)"
    ),
}


def _get_verify(insecure: bool) -> bool | str:
    if insecure:
        return False
    try:
        import certifi
        return certifi.where()
    except ImportError:
        print(
            "AVISO: falta 'certifi' -- si la descarga falla con "
            "CERTIFICATE_VERIFY_FAILED, instala 'certifi' o vuelve a "
            "correr con --insecure."
        )
        return True


def _intentar_url(url: str, *, insecure: bool) -> bytes | None:
    """Un único intento de GET -- separado de `_descargar`/
    `_descargar_con_fallback_de_año` para que ambas compartan la misma
    lógica de red (verificación SSL, User-Agent, comprobación de tamaño
    mínimo) sin duplicarla. Devuelve el contenido si parece un .xlsx
    real, o `None` si falla por cualquier motivo (ya impreso)."""
    try:
        import httpx
    except ImportError:
        print("Falta 'httpx' para la descarga automática. Instálalo con: pip install httpx --break-system-packages")
        return None

    try:
        resp = httpx.get(
            url, timeout=60, follow_redirects=True, verify=_get_verify(insecure),
            headers={"User-Agent": "Mozilla/5.0 (compatible; AmITraceable-TFG-research/1.0)"},
        )
        resp.raise_for_status()
    except Exception as e:
        print(f"  FALLO ({url}): {e!r}")
        return None

    if len(resp.content) < 1000:
        # Un .xlsx real pesa como mínimo varias decenas de KB -- una
        # respuesta de pocos bytes normalmente es una página de error
        # (404/redirección a login) disfrazada de 200 OK, no el fichero.
        print(f"  FALLO ({url}): la respuesta solo tiene {len(resp.content)} bytes -- probablemente no es el fichero real.")
        return None

    return resp.content


def _descargar(clave: str, destino: Path, *, insecure: bool = False) -> Path | None:
    """Para los ficheros SIN año en la URL (`_URLS_SIN_AÑO`) -- una única
    URL fija, sin nada que probar en bucle (ver el comentario de
    `_URLS_SIN_AÑO` para por qué no tiene sentido intentar variar el año
    aquí). Devuelve la ruta si funciona, `None` si falla -- en cuyo caso
    imprime el motivo exacto Y las instrucciones de navegación manual de
    `_INSTRUCCIONES_MANUALES`.

    IMPORTANTE: esto puede funcionar desde el ordenador de quien lo corre
    aunque Claude no pueda comprobarlo por su cuenta -- el bloqueo que ve
    Claude (robots.txt) es una política de la propia herramienta de
    fetch de Claude, no necesariamente un rechazo real del servidor ante
    una petición HTTP normal. Aun así, esto NO se ha probado de verdad
    contra la red real todavía -- la primera vez que se corra dirá si el
    supuesto era correcto."""
    url = _URLS_SIN_AÑO[clave]
    print(f"Intentando descargar '{clave}' de:\n  {url}")
    contenido = _intentar_url(url, insecure=insecure)
    if contenido is None:
        print(f"\n  Descarga manual para '{clave}':\n  {_INSTRUCCIONES_MANUALES[clave]}\n")
        return None

    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_bytes(contenido)
    print(f"  OK -- {len(contenido):,} bytes guardados en {destino}")
    return destino


def _descargar_con_fallback_de_año(clave: str, destino: Path, *, insecure: bool = False) -> Path | None:
    """Para los ficheros CON año en la URL (`_PLANTILLAS_URL_CON_AÑO`) --
    para cada año (de `hoy + _MARGEN_AÑOS_FUTUROS` hacia atrás hasta
    `_PRIMER_AÑO_CONOCIDO`), prueba TODAS las plantillas de esa clave en
    orden, y se queda con la PRIMERA combinación año+plantilla que
    responda -- así el script sigue funcionando en cursos futuros sin
    tocar código, y también se adapta solo si cambia el dominio/formato
    que funciona (ver comentario de `_PLANTILLAS_URL_CON_AÑO`).

    Pequeña pausa entre intentos (`_PAUSA_ENTRE_INTENTOS_SEG`) por si el
    'Server disconnected without sending a response' visto en la primera
    prueba real era una protección anti-bot ante peticiones seguidas
    demasiado rápidas, no un problema de URL -- no cuesta nada probarlo.

    Si la combinación que funciona es un fichero .px/CSV (no .xlsx), se
    guarda igualmente pero con un AVISO explícito: `procesar_historico_rama`
    solo sabe leer .xlsx por ahora -- pásamelo y le añado el parseo de
    CSV con el formato real delante, en vez de adivinarlo sin verlo."""
    import time
    from datetime import date as _date

    plantillas = _PLANTILLAS_URL_CON_AÑO[clave]
    año_mas_reciente_a_probar = _date.today().year + _MARGEN_AÑOS_FUTUROS
    print(f"Buscando '{clave}' probando años {año_mas_reciente_a_probar} -> {_PRIMER_AÑO_CONOCIDO}, {len(plantillas)} plantilla(s) por año:")

    primer_intento = True
    for año in range(año_mas_reciente_a_probar, _PRIMER_AÑO_CONOCIDO - 1, -1):
        for plantilla in plantillas:
            if not primer_intento:
                time.sleep(_PAUSA_ENTRE_INTENTOS_SEG)
            primer_intento = False

            url = plantilla.format(año=año)
            print(f"  Probando curso {año}: {url}")
            contenido = _intentar_url(url, insecure=insecure)
            if contenido is None:
                continue

            extension_real = ".xlsx" if url.endswith(".xlsx") else ".csv"
            destino_real = destino.with_suffix(extension_real)
            destino_real.parent.mkdir(parents=True, exist_ok=True)
            destino_real.write_bytes(contenido)
            print(f"  OK -- encontrado curso {año}, {len(contenido):,} bytes guardados en {destino_real}")
            if extension_real == ".csv":
                print(
                    "  AVISO: esto es un .px/CSV, no un .xlsx -- "
                    "`procesar_historico_rama` todavía no sabe leer este "
                    "formato (solo se ha implementado el parseo de .xlsx). "
                    "Pásame el fichero para añadir ese parseo con el "
                    "formato real delante -- mientras tanto, el método de "
                    "respaldo se sigue usando igual."
                )
                return None
            return destino_real

    print(f"\n  Ninguna combinación de año+plantilla entre {_PRIMER_AÑO_CONOCIDO} y {año_mas_reciente_a_probar} funcionó.")
    print(f"  Descarga manual para '{clave}':\n  {_INSTRUCCIONES_MANUALES[clave]}\n")
    return None


def _sin_acentos(texto: str) -> str:
    nfkd = unicodedata.normalize("NFKD", texto.lower())
    return "".join(c for c in nfkd if not unicodedata.combining(c))


# Palabras clave (sin acentos) por categoría de STUDIES_DISTRIBUTION --
# igual que en compute_studies_distribution.py (versión anterior de este
# cálculo), sin cambios en la clasificación en sí, solo en cómo se agrega
# el resultado final.
PALABRAS_CLAVE: dict[str, list[str]] = {
    "medicina": [r"\bmedicina\b"],
    "enfermeria": [r"\benfermeria\b"],
    "derecho": [r"\bderecho\b"],
    "ingenieria_informatica": [r"informatica", r"ingenieria.*computadores", r"ciencia.*datos"],
    "ingenieria_industrial": [r"ingenieria industrial", r"organizacion industrial"],
    "administracion_de_empresas": [r"administracion y direccion de empresas", r"\bade\b", r"administracion de empresas"],
    "psicologia": [r"psicologia"],
    "magisterio": [r"educacion infantil", r"educacion primaria", r"\bmagisterio\b", r"\bmaestro\b"],
    "arquitectura": [r"\barquitectura\b"],
    "farmacia": [r"\bfarmacia\b"],
    "biologia": [r"\bbiologia\b"],
    "periodismo": [r"periodismo"],
    "economia": [r"\beconomia\b"],
    "veterinaria": [r"veterinaria"],
}
_PATRONES_COMPILADOS = {c: [re.compile(p) for p in ps] for c, ps in PALABRAS_CLAVE.items()}

# A qué "Rama" oficial (las 5 categorías amplias del Ministerio)
# pertenece cada una de nuestras 14 categorías -- clasificación estándar
# ISCED/Ministerio de Universidades, NO una decisión nuestra: Psicología,
# Magisterio/Educación, Periodismo, Economía, ADE y Derecho son
# "Ciencias Sociales y Jurídicas"; Medicina/Enfermería/Farmacia/
# Veterinaria son "Ciencias de la Salud"; Informática/Industrial/
# Arquitectura son "Ingeniería y Arquitectura"; Biología es "Ciencias".
# Ninguna de nuestras 14 categorías cae en "Artes y Humanidades" -- esa
# rama no contribuye a STUDIES_DISTRIBUTION con el diccionario actual.
CATEGORIA_A_RAMA: dict[str, str] = {
    "medicina": "Ciencias de la Salud",
    "enfermeria": "Ciencias de la Salud",
    "farmacia": "Ciencias de la Salud",
    "veterinaria": "Ciencias de la Salud",
    "derecho": "Ciencias Sociales y Jurídicas",
    "administracion_de_empresas": "Ciencias Sociales y Jurídicas",
    "economia": "Ciencias Sociales y Jurídicas",
    "psicologia": "Ciencias Sociales y Jurídicas",
    "magisterio": "Ciencias Sociales y Jurídicas",
    "periodismo": "Ciencias Sociales y Jurídicas",
    "ingenieria_informatica": "Ingeniería y Arquitectura",
    "ingenieria_industrial": "Ingeniería y Arquitectura",
    "arquitectura": "Ingeniería y Arquitectura",
    "biologia": "Ciencias",
}


def _clasificar(titulacion: str) -> list[str]:
    texto = _sin_acentos(titulacion)
    return [c for c, patrones in _PATRONES_COMPILADOS.items() if any(p.search(texto) for p in patrones)]


def _valor_mas_reciente_de_fila(row: tuple, columna_inicio: int) -> float | None:
    num_años = (len(row) - columna_inicio) // 2
    for i in range(num_años):
        valor = row[columna_inicio + i * 2]
        if isinstance(valor, (int, float)):
            return float(valor)
    return None


def _encontrar_fila_cabecera(ws, columna_titulacion: int = 3, texto: str = "Titulación") -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and len(row) > columna_titulacion and row[columna_titulacion] == texto:
            return i
    raise ValueError(f"No se encontró la fila de cabecera ('{texto}' en columna {columna_titulacion}) en la hoja '{ws.title}'")


def procesar_titulacion(ruta: Path) -> tuple[dict[str, float], float, list[dict]]:
    """Lee un Excel de matriculados/egresados POR TITULACIÓN (formato
    ancho: CCAA, Universidad, Rama, Titulación, luego pares Valor/%Mujeres
    por año, más reciente primero). Devuelve (matriculados_por_categoria,
    total_general, filas_completas_para_exportar)."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    por_categoria: dict[str, float] = defaultdict(float)
    total = 0.0
    filas_completas: list[dict] = []

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        fila_cabecera = _encontrar_fila_cabecera(ws)
        for row in ws.iter_rows(min_row=fila_cabecera + 1, values_only=True):
            if not row or not row[3]:
                continue
            ccaa, universidad, rama, titulacion = row[0], row[1], row[2], row[3]
            valor = _valor_mas_reciente_de_fila(row, columna_inicio=4)
            if valor is None:
                continue
            total += valor
            categorias = _clasificar(titulacion)
            for categoria in categorias:
                por_categoria[categoria] += valor
            filas_completas.append({
                "hoja": nombre_hoja,
                "ccaa": ccaa,
                "universidad": universidad,
                "rama": rama,
                "titulacion": titulacion,
                "valor_mas_reciente": valor,
                "categorias_app": categorias,
            })

    return dict(por_categoria), total, filas_completas


def procesar_historico_rama(ruta: Path) -> dict[str, float]:
    """Lee el Excel histórico 'Total SUE por Rama' -- FORMATO REAL
    (confirmado contra el fichero real, ya no es un supuesto sin
    verificar): NO es una fila por año como se asumía antes de ver el
    fichero. Es una tabla ANCHA con dos dimensiones simultáneas:

    - COLUMNAS: 6 BLOQUES de ~39 columnas cada uno (un año 1985-1986...
      por columna DENTRO de cada bloque) -- el primer bloque es 'Total'
      (todas las ramas sumadas), y los otros 5 son cada rama. El nombre
      de cada bloque solo aparece en la fila de cabecera (fila 7), en la
      PRIMERA columna de ese bloque -- no repetido en cada columna.
    - FILAS: jerarquía anidada por indentación (Universidad > Nivel de
      estudio > Sexo), p. ej. 'Todas las universidades' > '    Total' >
      '        Ambos sexos'. La fila que necesitamos (universidad =
      'Todas las universidades' -- el agregado nacional, no una
      universidad concreta --, nivel = 'Total' -- Grado+Máster+Doctorado
      juntos, coherente con que este fichero es 'Total SUE' --, sexo =
      'Ambos sexos') es UNA sola fila, y esa misma fila vale para los 6
      bloques de columnas a la vez (las filas no se repiten por bloque,
      solo las columnas).

    BUG ANTERIOR (versión previa de esta función, corregido en esta
    sesión tras ver el fichero real): al buscar nombres de rama en las
    primeras 15 filas sin restringir columnas, encontraba la fila 7
    (cabecera de BLOQUE, una celda por bloque) y la trataba como si fuera
    una cabecera de UNA COLUMNA POR AÑO -- así que solo cogía la PRIMERA
    columna de cada bloque (el año 1985-1986) y la sumaba a través de
    TODAS las filas de la hoja (2136 filas, mezclando universidades,
    niveles y sexos sin distinguir) en vez de sumar los 39 años dentro de
    la fila correcta. De ahí el resultado sin sentido ('230 cursos
    sumados', cifras demasiado bajas)."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 1. Encontrar la fila de cabecera de BLOQUES (la que tiene "Total" y
    # los nombres de rama, cada uno en la PRIMERA columna de su bloque) y
    # construir los rangos [inicio, fin) de columnas de cada bloque.
    nombres_rama_validos = {
        "Ciencias de la Salud", "Ciencias Sociales y Jurídicas",
        "Ingeniería y Arquitectura", "Ciencias", "Artes y Humanidades",
    }
    fila_cabecera_bloques = None
    posiciones: list[tuple[int, str]] = []  # (columna 0-indexada, nombre de bloque)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        for j, celda in enumerate(row or []):
            if isinstance(celda, str) and celda.strip() in (nombres_rama_validos | {"Total"}):
                posiciones.append((j, celda.strip()))
        if len(posiciones) >= 2:  # "Total" + al menos una rama ya es suficiente para confirmar que es esta fila
            fila_cabecera_bloques = i
            break
        posiciones = []

    if fila_cabecera_bloques is None or len(posiciones) < 6:
        raise ValueError(
            f"No se encontraron los 6 bloques de columna esperados ('Total' + 5 ramas) "
            f"en '{ruta.name}' -- el formato de este fichero no es el esperado "
            f"(se encontraron {len(posiciones)}). Revisa las primeras 15 filas a mano "
            f"y ajusta `procesar_historico_rama` con el formato real."
        )

    posiciones.sort(key=lambda p: p[0])
    rangos_por_bloque: dict[str, tuple[int, int]] = {}
    for idx, (col_inicio, nombre) in enumerate(posiciones):
        col_fin = posiciones[idx + 1][0] if idx + 1 < len(posiciones) else ws.max_column
        rangos_por_bloque[nombre] = (col_inicio, col_fin)

    # 2. Encontrar la fila objetivo: 'Todas las universidades' (agregado
    # nacional) -> dentro de ella, 'Total' (nivel, no Grado/Máster suelto)
    # -> dentro de ese, 'Ambos sexos'. Se busca secuencialmente en vez de
    # asumir números de fila fijos, por si la maquetación cambia de un
    # curso a otro.
    fila_universidad = fila_nivel = fila_sexo = None
    for i, row in enumerate(ws.iter_rows(min_row=fila_cabecera_bloques + 1, values_only=True), start=fila_cabecera_bloques + 1):
        etiqueta = (row[0] or "").strip() if row and row[0] else ""
        if fila_universidad is None:
            if etiqueta == "Todas las universidades":
                fila_universidad = i
            continue
        if fila_nivel is None:
            if etiqueta == "Total":
                fila_nivel = i
            continue
        if fila_sexo is None:
            if etiqueta == "Ambos sexos":
                fila_sexo = i
            break

    if fila_sexo is None:
        raise ValueError(
            f"No se encontró la fila 'Todas las universidades' > 'Total' > 'Ambos sexos' "
            f"en '{ruta.name}' -- revisa la jerarquía de filas a mano y ajusta "
            f"`procesar_historico_rama` con el formato real."
        )

    fila_valores = list(ws.iter_rows(min_row=fila_sexo, max_row=fila_sexo, values_only=True))[0]

    # 3. Sumar, DENTRO de esa única fila, las columnas de cada bloque de
    # rama (39 años cada uno) -- así se obtiene el acumulado histórico
    # real desde 1985-1986, no una mezcla de filas no relacionadas.
    acumulado: dict[str, float] = {}
    for nombre, (col_inicio, col_fin) in rangos_por_bloque.items():
        if nombre == "Total":
            continue  # es la suma de las 5 ramas, no una rama en sí -- se usa solo como comprobación más abajo
        valores = [v for v in fila_valores[col_inicio:col_fin] if isinstance(v, (int, float))]
        acumulado[nombre] = sum(valores)

    # Comprobación de consistencia: la suma de las 5 ramas debería
    # cuadrar (aprox.) con el bloque "Total" de la misma fila -- si no
    # cuadra, algo del parseo de rangos de columna está mal.
    col_inicio_total, col_fin_total = rangos_por_bloque["Total"]
    valores_total = [v for v in fila_valores[col_inicio_total:col_fin_total] if isinstance(v, (int, float))]
    suma_total_bloque = sum(valores_total)
    suma_5_ramas = sum(acumulado.values())
    if suma_total_bloque and abs(suma_5_ramas - suma_total_bloque) / suma_total_bloque > 0.02:
        print(
            f"  AVISO: la suma de las 5 ramas ({suma_5_ramas:,.0f}) no cuadra con el "
            f"bloque 'Total' de la misma fila ({suma_total_bloque:,.0f}) -- diferencia "
            f"de {abs(suma_5_ramas - suma_total_bloque) / suma_total_bloque:.1%}. Revisar "
            f"los rangos de columna antes de confiar en el resultado."
        )

    print(
        f"  [{ruta.name}] fila usada: 'Todas las universidades' (fila {fila_universidad}) > "
        f"'Total' (fila {fila_nivel}) > 'Ambos sexos' (fila {fila_sexo}); "
        f"{col_fin_total - col_inicio_total} cursos por bloque; "
        f"comprobación 5 ramas vs Total: {suma_5_ramas:,.0f} vs {suma_total_bloque:,.0f}"
    )
    return acumulado


def _poblacion_25_64() -> float:
    """Población 25-64 REAL (no un porcentaje asumido a mano), calculada
    a partir de TOTAL_POPULATION_ES x suma de los tramos 25-29..60-64 de
    AGE_DISTRIBUTION_5Y -- ambas constantes ya existentes en
    ine_reference.py, importadas directamente en vez de copiadas, para
    que si se actualizan ahí este cálculo las recoja automáticamente."""
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from app.data import ine_reference  # noqa: E402
    tramos = ["25-29", "30-34", "35-39", "40-44", "45-49", "50-54", "55-59", "60-64"]
    fraccion = sum(ine_reference.AGE_DISTRIBUTION_5Y[t] for t in tramos)
    return ine_reference.TOTAL_POPULATION_ES * fraccion


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "--matriculados", type=Path, default=None,
        help="Excel de Matriculados por titulación (2015-2024) -- si no se da, se intenta descargar solo",
    )
    parser.add_argument(
        "--egresados", type=Path, default=None,
        help="Excel de Egresados por titulación (2015-2023) -- si no se da, se intenta descargar solo",
    )
    parser.add_argument(
        "--egresados-historico-rama", type=Path, default=None,
        help="Excel histórico 'Egresados Total SUE por Rama' desde 1984 -- si no se da, se intenta "
        "descargar solo; si la descarga también falla, se usa el método de respaldo (40%% asumido) "
        "con aviso explícito -- a diferencia de matriculados/egresados, esta NO es obligatoria.",
    )
    parser.add_argument(
        "--no-download", action="store_true",
        help="No intentar descargar nada automáticamente -- falla con instrucciones manuales si "
        "falta algún fichero en vez de intentar bajarlo",
    )
    parser.add_argument("--insecure", action="store_true", help="Desactivar verificación SSL al descargar (diagnóstico/último recurso)")
    parser.add_argument("--cache-dir", type=Path, default=_CACHE_DIR, help=f"Dónde guardar los ficheros descargados (por defecto: {_CACHE_DIR})")
    parser.add_argument("--apply", action="store_true", help="Escribe el resultado en ine_reference.py (si no, solo lo imprime)")
    parser.add_argument("--yes", action="store_true", help="No pedir confirmación antes de escribir")
    parser.add_argument(
        "--export-json", type=Path, default=_JSON_EXPORT_PATH,
        help=f"Ruta del JSON con el detalle completo por universidad/titulación (por defecto: {_JSON_EXPORT_PATH})",
    )
    parser.add_argument("--no-export-json", action="store_true", help="No exportar el detalle completo por universidad")
    args = parser.parse_args()

    # --- Resolver las rutas de entrada: usar la que se haya pasado a
    # mano, o intentar descargarla automáticamente si no. matriculados y
    # egresados son OBLIGATORIOS (si fallan los dos, no hay nada que
    # calcular); egresados_historico_rama es OPCIONAL (si falla, sigue
    # el método de respaldo, ya implementado más abajo). ---
    ruta_matriculados = args.matriculados
    if ruta_matriculados is None and not args.no_download:
        ruta_matriculados = _descargar("matriculados", args.cache_dir / "MatriculadosTitulacion.xlsx", insecure=args.insecure)
    if ruta_matriculados is None:
        print("\nERROR: no hay fichero de MATRICULADOS (ni pasado a mano ni descargado) -- no se puede continuar.")
        sys.exit(1)

    ruta_egresados = args.egresados
    if ruta_egresados is None and not args.no_download:
        ruta_egresados = _descargar("egresados", args.cache_dir / "EgresadosTitulacion.xlsx", insecure=args.insecure)
    if ruta_egresados is None:
        print("\nERROR: no hay fichero de EGRESADOS (ni pasado a mano ni descargado) -- no se puede continuar.")
        sys.exit(1)

    ruta_historico = args.egresados_historico_rama
    if ruta_historico is None and not args.no_download:
        ruta_historico = _descargar_con_fallback_de_año("egresados_historico_rama", args.cache_dir / "EgresadosHistoricoRama.xlsx", insecure=args.insecure)
    if ruta_historico is None:
        print(
            "\nAVISO: no hay fichero histórico por rama (ni pasado a mano ni "
            "descargado) -- se seguirá con el método de respaldo más abajo. "
            "Esto NO es un error fatal, solo reduce la precisión del resultado."
        )

    print(f"\n=== Procesando MATRICULADOS: {ruta_matriculados} ===")
    mat_por_categoria, mat_total, mat_filas = procesar_titulacion(ruta_matriculados)
    print(f"  Total matriculados: {mat_total:,.0f}")

    print(f"\n=== Procesando EGRESADOS: {ruta_egresados} ===")
    egr_por_categoria, egr_total, egr_filas = procesar_titulacion(ruta_egresados)
    print(f"  Total egresados (suma de años disponibles en el fichero, NO acumulado histórico): {egr_total:,.0f}")

    if not args.no_export_json:
        detalle_completo = {
            "matriculados": mat_filas,
            "egresados": egr_filas,
            "nota": (
                "Detalle completo por universidad/titulación/año, exportado por "
                "update_studies_distribution.py. Pensado para análisis futuros "
                "que sepan la universidad y el año concretos de alguien (p. ej. "
                "'estudió en la Universidad de Oviedo sobre 2015') -- "
                "STUDIES_DISTRIBUTION en ine_reference.py solo usa el agregado "
                "nacional, no este detalle."
            ),
        }
        args.export_json.parent.mkdir(parents=True, exist_ok=True)
        args.export_json.write_text(json.dumps(detalle_completo, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\nDetalle completo exportado a {args.export_json} ({len(mat_filas) + len(egr_filas)} filas).")

    # Reparto reciente dentro de cada rama, usando EGRESADOS como base
    # preferente (más adecuado conceptualmente que matriculados para
    # aproximar "personas tituladas", ver docstring) -- si una categoría
    # no tiene egresados pero sí matriculados (programa nuevo), se usa
    # matriculados como respaldo para esa categoría concreta.
    reparto_reciente: dict[str, float] = {}
    for categoria in PALABRAS_CLAVE:
        valor = egr_por_categoria.get(categoria)
        if valor is None or valor == 0:
            valor = mat_por_categoria.get(categoria, 0.0)
        reparto_reciente[categoria] = valor

    resultado_final: dict[str, float] | None = None

    if ruta_historico is not None:
        print(f"\n=== Procesando HISTÓRICO POR RAMA: {ruta_historico} ===")
        try:
            total_historico_por_rama = procesar_historico_rama(ruta_historico)
        except Exception as e:
            print(
                f"  FALLO al parsear el histórico ({e!r}) -- el formato real "
                "de este fichero no coincide con lo que asumía "
                "`procesar_historico_rama` (ver el docstring de esa función). "
                "Cayendo al método de respaldo en vez de fallar del todo."
            )
            total_historico_por_rama = None

        if total_historico_por_rama:
            print(f"  Totales acumulados por rama: {total_historico_por_rama}")
            print(
                "\nUsando el MÉTODO DE DOS PASOS (total histórico por rama x "
                "reparto reciente dentro de la rama) -- ver docstring de este "
                "script para el razonamiento completo."
            )
            total_estimado: dict[str, float] = {}
            for categoria in PALABRAS_CLAVE:
                rama = CATEGORIA_A_RAMA.get(categoria)
                if rama is None:
                    total_estimado[categoria] = 0.0
                    continue
                total_rama_historico = total_historico_por_rama.get(rama, 0.0)
                suma_reparto_rama = sum(reparto_reciente[c] for c, r in CATEGORIA_A_RAMA.items() if r == rama)
                if suma_reparto_rama == 0:
                    total_estimado[categoria] = 0.0
                    continue
                proporcion_dentro_de_rama = reparto_reciente[categoria] / suma_reparto_rama
                total_estimado[categoria] = total_rama_historico * proporcion_dentro_de_rama

            poblacion = _poblacion_25_64()
            print(f"  Población 25-64 usada como denominador (calculada, no asumida): {poblacion:,.0f}")
            resultado_final = {c: total_estimado[c] / poblacion for c in PALABRAS_CLAVE}

    if resultado_final is None:
        # Se llega aquí tanto si no se dio --egresados-historico-rama como
        # si se dio pero el parseo falló (ver el try/except de arriba) --
        # en ambos casos, mismo método de respaldo.
        print(
            "\nAVISO: usando el MÉTODO DE RESPALDO (reparto reciente sobre "
            "MATRICULADOS -- no egresados, para que numerador y denominador "
            "usen la misma base -- x 40% de población 25-64 con estudios "
            "superiores, supuesto SIN verificar) en vez del método de dos "
            "pasos con datos históricos reales. Los resultados serán menos "
            "fiables que si se consigue el fichero histórico."
        )
        resultado_final = {
            c: (mat_por_categoria.get(c, 0.0) / mat_total if mat_total else 0.0) * 0.40
            for c in PALABRAS_CLAVE
        }

    print("\n" + "=" * 78)
    print("RESULTADO: STUDIES_DISTRIBUTION")
    print("=" * 78)
    resultado_para_aplicar: dict[str, float] = {}
    for categoria in PALABRAS_CLAVE:
        clave_app = categoria.replace("_", " ")
        valor = resultado_final[categoria]
        resultado_para_aplicar[clave_app] = valor
        print(f'    "{clave_app}": {valor:.4f},')

    if args.apply:
        lines = _uir._INE_REFERENCE_PATH.read_text(encoding="utf-8").splitlines(keepends=True)
        block = _uir._locate_block(lines, lambda line: line.startswith("STUDIES_DISTRIBUTION = {"))
        if block is None:
            print("\nERROR: no se encontró 'STUDIES_DISTRIBUTION = { ... }' en ine_reference.py -- no se ha tocado nada.")
            sys.exit(1)
        start, end = block
        changes = _uir._apply_float_block(lines, start, end, resultado_para_aplicar, ndigits=4)
        if changes:
            _uir._update_last_verified(lines, "STUDIES_DISTRIBUTION")
        _uir._confirm_and_write(changes, "STUDIES_DISTRIBUTION", lines, auto_confirm=args.yes)
    else:
        print("\n(no se ha pasado --apply -- esto es solo una previsualización, no se ha escrito nada)")


if __name__ == "__main__":
    main()
